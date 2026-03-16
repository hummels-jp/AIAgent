import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import io

# 设置输出编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 创建新工作簿
wb = openpyxl.Workbook()

# 样式定义
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='Arial', size=14, bold=True, color='000000')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 工作表1: 详细对比表
ws1 = wb.active
ws1.title = '詳細比較テーブル'

# 标题
ws1.merge_cells('A1:E1')
ws1['A1'] = 'OpenDriveとOSMデータ形式の詳細比較'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

# 列标题
headers = ['特徴', 'kawasaki.osm', 'kawasaki.net.xml', 'kawasaki.xodr (SUMO)', 'dgm_kawasaki.xodr (DGM)']
for col, header in enumerate(headers, start=1):
    cell = ws1.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# 数据
data = [
    ['主な用途', '実地理データ保存', '交通流シミュレーション', '運転シミュレーション（ADASテスト）', '高精度地図/可視化'],
    ['座標系', 'WGS84 (lat/lon)', 'UTM Zone 54', 'UTM + offset', '直接地理坐标 (大範囲)'],
    ['幾何精度', 'ノード座標', '車線形状', '高精度道路幾何', '超高精度（16桁小数）'],
    ['ファイルサイズ', '1,860 KB', '1,118 KB', '3,269 KB', '8,404 KB'],
    ['行数', '23,828', '10,680', '72,885', '~395,000'],
    ['データ標準化', 'OSM仕様', 'SUMO仕様', 'OpenDRIVE 1.4標準', 'OpenDRIVE 1.5標準'],
    ['応用シナリオ', '地図編集', '交通管理', '自動運転テスト', '高精度シミュレーション/可視化'],
    ['道路数', '-', '-', '~471', '~476'],
    ['交差点数', '-', '-', '~217', '~285'],
    ['速度制限', 'あり（タグ）', 'あり（属性）', 'あり（曲線）', 'なし'],
    ['surface要素', '-', '-', 'なし', '有'],
    ['userData拡張', '-', '-', 'なし', '有 (viStyleDef)'],
    ['', '', '', '', ''],
    ['メタデータの豊富さ', '⭐⭐⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐'],
    ['シミュレーション精度', '-', '中等', '高精度', '超高精度']
]

for row_idx, row_data in enumerate(data, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx == 1 else center_align
        cell.border = border
        # 设置列宽
        if col_idx == 1:
            ws1.column_dimensions['A'].width = 25
        else:
            ws1.column_dimensions[chr(64 + col_idx)].width = 30

# 工作表2: 主要差异总结
ws2 = wb.create_sheet('主な違いの要約')

# 标题
ws2.merge_cells('A1:G1')
ws2['A1'] = 'OpenDriveとOSMデータ形式の主要な違い'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

# 列标题
headers2 = ['分類', '違い', 'OpenDriveの規定', 'OSMの規定', 'kawasaki.xodr (SUMO)', 'dgm_kawasaki.xodr (DGM)', '影響']
for col, header in enumerate(headers2, start=1):
    cell = ws2.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# 数据分组1: 基本格式
for i in range(3, 6):
    ws2[f'A{i}'] = '基本格式'
    ws2[f'A{i}'].alignment = center_align
    ws2[f'A{i}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data1 = [
    ['Headerバージョン', '', '', '', 'revMinor=4 (1.4)', 'revMinor=5 (1.5)', 'DGMはより多くの拡張属性をサポート'],
    ['座標系统', '', '', '', 'UTM + offset', '直接地理坐标 (~1554万/421万)', '統一には座標変換が必要'],
    ['数值フォーマット', '', '', '', '常规小数 (107.64497470)', '科学计数法 (3.4112208657352156e+01)', 'DGMの精度が高いが、特別な処理が必要']
]

for row_idx, row_data in enumerate(data1, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        if col_idx == 1 and row_idx == 3:  # 跳过合并的标题单元格
            continue
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组2: 内容
ws2.merge_cells('A6:A10')
ws2['A6'] = 'コンテンツ'
ws2['A6'].alignment = center_align
ws2['A6'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data2 = [
    ['車線幅', '', '', '', '3.20米', '3.75米', '道路幅の違い'],
    ['車線マーキング', '', '', '', 'width=0.13, なしheight', 'width=0.15, height=0.02, laneChange', 'DGMがより詳細'],
    ['速度情報', '', '', '', '有 (5.56, 11.11, 13.89 m/s)', 'なし', 'SUMOはシミュレーションにより適している'],
    ['rule属性', '', '', '', 'なし', '所有道路rule="RHT"', 'DGMは交通ルールを明確に定義'],
    ['surface要素', '', '', '', 'なし', '每个道路都有', 'DGMは路面属性をサポート']
]

for row_idx, row_data in enumerate(data2, start=6):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组3: 道路几何
ws2.merge_cells('A11:A16')
ws2['A11'] = '道路幾何'
ws2['A11'].alignment = center_align
ws2['A11'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data3 = [
    ['道路基準線', '', '〇', '', '', '', ''],
    ['曲率（line / arc / clothoid）', '', '〇', '', '', '', ''],
    ['道路長', '', '〇', '', '', '', ''],
    ['標高（elevation）', '', '〇', '', '', '', ''],
    ['横断勾配（superelevation）', '', '〇', '', '', '', ''],
    ['横断形状（lateral profile）', '', '〇', '', '', '', '']
]

for row_idx, row_data in enumerate(data3, start=11):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组4: 车道结构
ws2.merge_cells('A17:A24')
ws2['A17'] = 'レーン構造'
ws2['A17'].alignment = center_align
ws2['A17'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data4 = [
    ['レーンID', '', '〇', '', '', '', ''],
    ['レーン幅', '', '〇', '', '', '', ''],
    ['レーンタイプ', '', '〇', '', '', '', ''],
    ['レーン接続関係', '', '〇', '', '', '', ''],
    ['レーン速度制限', '', '〇', '', '', '', ''],
    ['レーンアクセス制限', '', '〇', '', '', '', ''],
    ['区画線（実線 / 破線）', '', '〇', '', '', '', ''],
    ['色 / 幅 / 車線変更可否 / パターン', '', '〇', '', '', '', '']
]

for row_idx, row_data in enumerate(data4, start=17):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组5: 交通标识
ws2.merge_cells('A25:A30')
ws2['A25'] = '交通標識・信号'
ws2['A25'].alignment = center_align
ws2['A25'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data5 = [
    ['信号機', '', '〇', '', '', '', ''],
    ['一時停止標識', '', '〇', '', '', '', ''],
    ['制限速度標識', '', '〇', '', '', '', ''],
    ['進入禁止', '', '〇', '', '', '', ''],
    ['車線別制限', '', '〇', '', '', '', ''],
    ['警告標識', '', '〇', '', '', '', '']
]

for row_idx, row_data in enumerate(data5, start=25):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组6: 道路对象
ws2.merge_cells('A31:A39')
ws2['A31'] = '道路オブジェクト'
ws2['A31'].alignment = center_align
ws2['A31'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data6 = [
    ['停止線', '', '〇', '', '', '', ''],
    ['横断歩道', '', '〇', '', '', '', ''],
    ['路面ペイント', '', '〇', '', '', '', ''],
    ['ガードレール', '', '〇', '', '', '', ''],
    ['ポール', '', '〇', '', '', '', ''],
    ['路肩縁石', '', '〇', '', '', '', ''],
    ['建物', '', '〇', '', '', '', ''],
    ['駐車スペース', '', '〇', '', '', '', ''],
    ['バリア / トンネル / 橋', '', '〇', '', '', '', '']
]

for row_idx, row_data in enumerate(data6, start=31):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组7: 交差点与连接
ws2.merge_cells('A40:A42')
ws2['A40'] = '交差点構造'
ws2['A40'].alignment = center_align
ws2['A40'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data7 = [
    ['接続道路 / turning relation', '', '〇', '', '', '', ''],
    ['lane connection', '', '〇', '', '', '', ''],
    ['路面特性（摩擦係数 / 材質 / roughness）', '', '〇', '', '', '', '']
]

for row_idx, row_data in enumerate(data7, start=40):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 数据分组8: 扩展
ws2.merge_cells('A43:A45')
ws2['A43'] = '拡張機能'
ws2['A43'].alignment = center_align
ws2['A43'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

data8 = [
    ['userData/viStyleDef', '', '', '', 'なし', '每个车道都有', 'DGMはカスタム拡張をサポート'],
    ['適用シナリオ', '', '', '', 'SUMO仿真、简单转换', '高精度地图、可视化、专业仿真', '要件に応じて選択'],
    ['データ精度', '', '', '', '高精度', '超高精度', '用途による']
]

for row_idx, row_data in enumerate(data8, start=43):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx <= 2 else center_align
        cell.border = border

# 设置列宽
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 25
ws2.column_dimensions['G'].width = 40

# 工作表3: 应用建议
ws3 = wb.create_sheet('応用シナリオ')

# 标题
ws3.merge_cells('A1:E1')
ws3['A1'] = 'OpenDriveとOSMの応用シナリオと推奨事項'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

# 列标题
headers3 = ['シナリオ', 'OSM推奨度', 'SUMO OpenDrive推奨度', 'DGM OpenDrive推奨度', '推奨形式']
for col, header in enumerate(headers3, start=1):
    cell = ws3.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# 数据
data3_content = [
    ['リアルタイムナビゲーション', '⭐⭐⭐⭐⭐', '⭐⭐', '⭐⭐', 'OSM'],
    ['地図編集・更新', '⭐⭐⭐⭐⭐', '⭐⭐', '⭐⭐', 'OSM'],
    ['交通流シミュレーション', '⭐⭐⭐', '⭐⭐⭐⭐⭐', '⭐⭐⭐', 'SUMO OpenDrive'],
    ['自動運転テスト', '⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM OpenDrive'],
    ['高精度可視化', '⭐⭐', '⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM OpenDrive'],
    ['車両ダイナミクス', '⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM OpenDrive'],
    ['ADAS機能検証', '⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM OpenDrive'],
    ['ロボットナビゲーション', '⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐⭐', 'DGM OpenDrive'],
    ['物流最適化', '⭐⭐⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐', 'OSM'],
    ['都市計画', '⭐⭐⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐', 'OSM'],
    ['軽量アプリケーション', '⭐⭐⭐⭐⭐', '⭐⭐', '⭐⭐', 'OSM'],
    ['プロフェッショナルシミュレーション', '⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM OpenDrive']
]

for row_idx, row_data in enumerate(data3_content, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws3.cell(row_idx, col_idx, value)
        cell.alignment = center_align
        cell.border = border

# 设置列宽
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 25
ws3.column_dimensions['E'].width = 15

# 工作表4: 转换策略
ws4 = wb.create_sheet('変換戦略')

# 标题
ws4.merge_cells('A1:C1')
ws4['A1'] = 'データ形式間の変換戦略'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

# 列标题
headers4 = ['変換方向', '主な課題', '推奨アプローチ']
for col, header in enumerate(headers4, start=1):
    cell = ws4.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# 数据
data4_content = [
    ['OSM → OpenDrive', 
     '幾何精度の補強、車線情報の補完、高度データの追加、交通規則の変換',
     'SUMO netconvertを使用、追加データ（DEM等）の統合、手動修正'],
    ['OpenDrive → OSM',
     '幾何の簡略化、階層構造のフラット化、拡張属性のマッピング',
     '曲線のサンプリング、OSMタグへのマッピング、精度の検証'],
    ['SUMO → DGM',
     'バージョンのアップグレード、詳細属性の追加',
     'SUMO出力を基礎、追加メタデータの付与'],
    ['DGM → SUMO',
     '不要な詳細の削減、互換性の維持',
     '主要属性のみ保持、変換スクリプトの使用'],
    ['相互運用性',
     '座標変換、単位変換、データ検証',
     '標準化ツールの使用、品質管理プロセスの導入']
]

for row_idx, row_data in enumerate(data4_content, start=3):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws4.cell(row_idx, col_idx, value)
        cell.alignment = left_align if col_idx in [1, 2] else center_align
        cell.border = border

# 设置列宽
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 40

# 保存文件
output_file = 'D:/huqianqian_git/AIAgent/20260313_OSM_SUMO_OpenDrive/2026_03_16_コンテンツ比較表_更新版.xlsx'
wb.save(output_file)

print(f'Excel文件已成功创建: {output_file}')
print('工作表包含:')
print('1. 詳細比較テーブル - 详细对比表')
print('2. 主な違いの要約 - 主要差异总结')
print('3. 応用シナリオ - 应用场景')
print('4. 変換戦略 - 转换策略')
