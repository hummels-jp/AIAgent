import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建新工作簿
wb = Workbook()
ws1 = wb.active
ws1.title = '詳細比較テーブル'

# 样式定义
title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 工作表1: 详细对比
ws1['A1'] = 'OpenDriveとOSMデータ形式の詳細比較'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

# 标题行
headers = ['特徴', 'kawasaki.osm', 'kawasaki.net.xml', 'kawasaki.xodr (SUMO)', 'dgm_kawasaki.xodr (DGM)']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# 数据行
data = [
    ['主な用途', '実地理データ保存', '交通流シミュレーション', '運転シミュレーション（ADASテスト）', '高精度地図/可視化'],
    ['座標系', 'WGS84 (lat/lon)', 'UTM Zone 54', 'UTM + offset', '直接地理坐标 (大範囲)'],
    ['幾何精度', 'ノード座標', '車線形状', '高精度道路幾何', '超高精度（16桁小数）'],
    ['ファイルサイズ', '1,860 KB', '1,118 KB', '3,269 KB', '8,404 KB'],
    ['行数', '23,828', '10,680', '72,885', '~395,000'],
    ['データ標準化', 'OSM仕様', 'SUMO仕様', 'OpenDRIVE 1.4標準', 'OpenDRIVE 1.5標準'],
    ['応用シナリオ', '地図編集', '交通管理', '自動運転テスト', '高精度シミュレーション/可視化'],
    ['道路数', '-', '-', '~471', '~476'],
    ['交差点数', '-', '-', '~217', '~285'],
    ['速度制限', 'あり（タグ）', 'あり（属性）', 'あり（曲線）', 'なし'],
    ['surface要素', '-', '-', 'なし', '有'],
    ['userData拡張', '-', '-', 'なし', '有 (viStyleDef)'],
    ['メタデータの豊富さ', '⭐⭐⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐'],
    ['シミュレーション精度', '-', '中等', '高精度', '超高精度']
]

for row_idx, row_data in enumerate(data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row_idx, col_idx, value)
        cell.alignment = center_align
        cell.border = border

# 列宽
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 30
ws1.column_dimensions['E'].width = 30

# 工作表2: 主要差异
ws2 = wb.create_sheet('主な違いの要約')
ws2['A1'] = 'OpenDriveとOSMの主要な違い'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

headers2 = ['分類', '項目', 'OpenDrive', 'OSM', 'SUMOバージョン', 'DGMバージョン', '影響']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

data2 = [
    ['基本格式', 'Headerバージョン', 'revMinor=4/5', '-', 'revMinor=4 (1.4)', 'revMinor=5 (1.5)', 'DGMはより多くの拡張'],
    ['', '座標系统', 'UTM/地理坐标', 'WGS84', 'UTM + offset', '直接地理坐标', '変換が必要'],
    ['', '数値フォーマット', '小数/科学記法', '小数', '常规小数', '科学記法', 'DGM精度が高い'],
    ['コンテンツ', '車線幅', '3.20-3.75m', '3.75m', '3.20米', '3.75米', '標準が異なる'],
    ['', '車線マーキング', '詳細あり', '基本あり', 'width=0.13', 'width=0.15, height=0.02', 'DGMが詳細'],
    ['', '速度情報', 'あり', 'あり', 'あり (5.56-13.89)', 'なし', 'SUMO向き'],
    ['', 'rule属性', 'あり', '-', 'なし', 'rule="RHT"', 'DGMが明確'],
    ['', 'surface要素', 'あり', '-', 'なし', 'あり', 'DGMが対応'],
    ['道路幾何', 'Reference line', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '曲率（line/arc/clothoid）', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '道路長', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '標高（elevation）', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '横断勾配', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '横断形状', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['レーン構造', 'レーンID', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'レーン幅', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'レーンタイプ', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'レーン接続', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '速度制限', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'アクセス制限', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['区画線', '実線/破線', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '色/幅/車線変更', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'パターン/複合線', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['交通標識・信号', '信号機', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '停止標識', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '制限速度標識', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '進入禁止', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '車線別制限', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '警告標識', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['道路オブジェクト', '停止線', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '横断歩道', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '路面ペイント', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'ガードレール/ポール', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '路肩縁石/建物', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '駐車スペース', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'バリア/トンネル/橋', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['交差点構造', '接続道路', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'turning relation', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'lane connection', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['路面特性', '摩擦係数', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', '路面材質', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['', 'roughness', '〇', '-', '〇', '〇', 'OpenDrive標準'],
    ['拡張機能', 'userData', 'あり', '-', 'なし', 'あり (viStyleDef)', 'DGM対応'],
    ['応用シナリオ', '適用', '-', '-', 'SUMOシミュレーション', '高精度シミュレーション']
]

for row_idx, row_data in enumerate(data2, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row_idx, col_idx, value)
        cell.alignment = center_align
        cell.border = border

# 列宽
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 20
ws2.column_dimensions['F'].width = 25
ws2.column_dimensions['G'].width = 25

# 工作表3: 应用场景
ws3 = wb.create_sheet('応用シナリオ')
ws3['A1'] = 'OpenDriveとOSMの応用シナリオ'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

headers3 = ['シナリオ', 'OSM', 'OpenDrive (SUMO)', 'OpenDrive (DGM)', '推奨']
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

data3 = [
    ['リアルタイムナビゲーション', '⭐⭐⭐⭐⭐', '⭐⭐', '⭐⭐', 'OSM'],
    ['地図編集・更新', '⭐⭐⭐⭐⭐', '⭐⭐', '⭐⭐', 'OSM'],
    ['交通流シミュレーション', '⭐⭐⭐', '⭐⭐⭐⭐⭐', '⭐⭐⭐', 'SUMO'],
    ['自動運転テスト', '⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM'],
    ['高精度可視化', '⭐⭐', '⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM'],
    ['車両ダイナミクス', '⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM'],
    ['ADAS検証', '⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM'],
    ['ロボットナビゲーション', '⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐⭐', 'DGM'],
    ['物流最適化', '⭐⭐⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐', 'OSM'],
    ['都市計画', '⭐⭐⭐⭐⭐', '⭐⭐⭐', '⭐⭐⭐', 'OSM'],
    ['軽量アプリ', '⭐⭐⭐⭐⭐', '⭐⭐', '⭐⭐', 'OSM'],
    ['プロフェッショナルシミュレーション', '⭐⭐', '⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'DGM']
]

for row_idx, row_data in enumerate(data3, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row_idx, col_idx, value)
        cell.alignment = center_align
        cell.border = border

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 12

# 工作表4: 转换建议
ws4 = wb.create_sheet('変換アドバイス')
ws4['A1'] = 'データ形式変換のアドバイス'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

headers4 = ['変換方向', '課題', '推奨アプローチ']
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(2, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

data4 = [
    ['OSM → OpenDrive', '幾何補強、情報補完、高度追加、規則変換', 'SUMO netconvert + DEM統合 + 手動修正'],
    ['OpenDrive → OSM', '幾何簡略化、フラット化、属性マッピング', '曲線サンプリング + タグマッピング + 検証'],
    ['SUMO → DGM', 'バージョンアップ、詳細属性追加', 'SUMO出力基礎 + メタデータ追加'],
    ['DGM → SUMO', '詳細削減、互換性維持', '主要属性保持 + 変換スクリプト'],
    ['相互運用性', '座標・単位変換、検証', '標準ツール使用、品質管理']
]

for row_idx, row_data in enumerate(data4, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row_idx, col_idx, value)
        cell.alignment = center_align if col_idx == 3 else left_align
        cell.border = border

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 40

# 保存
output_file = 'D:/huqianqian_git/AIAgent/20260313_OSM_SUMO_OpenDrive/2026_03_16_コンテンツ比較表_完全版.xlsx'
wb.save(output_file)
print(f'成功创建: {output_file}')
